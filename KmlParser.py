import xml.sax
import openpyxl

class KmlHandler(xml.sax.ContentHandler):
    def __init__(self):
        self.content = ""
        self.name = ""
        self.site_coordinate = ""
        self.site_list = []
        self.line_list = []
        self.in_placemark = 0
        self.is_point = 0
        self.is_line = 0

    def write_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell("A1").value = "Name"
        ws.cell("B1").value = "Longtitude 1"
        ws.cell("C1").value = "Latitude 1"
        ws.cell("D1").value = "Longtitude 2"
        ws.cell("E1").value = "Latitude 2"
        for site in self.site_list:
            ws.append(site)
        for line in self.line_list:
            ws.append(line)
        wb.save("test.xlsx")

    def startElement(self, tag, attributes):
        self.content = tag
        # print("Start tag:" + tag)
        if tag == "Placemark":
            self.in_placemark = 1
        elif tag == "Point":
            self.is_point = 1
        elif tag == "LineString":
            self.is_line = 1
        else:
            return

    def endElement(self, tag):
        # print("End tag:" + tag)
        if tag == "Placemark":
            co_list = self.site_coordinate.split(",")
            if self.is_point == 1 and len(co_list) >= 3:
                self.site_list.append([self.name, co_list[0], co_list[1]])
            elif self.is_line == 1:
                site_co_list = self.site_coordinate.split(" ")
                site1_coordinate = site_co_list[0].split(",")
                site2_coordinate = site_co_list[1].split(",")
                self.line_list.append([self.name, site1_coordinate[0], site1_coordinate[1], site2_coordinate[0], site2_coordinate[1]])
            self.name = ""
            self.site_coordinate = ""
            self.in_placemark = 0
            self.is_point = 0
            self.is_line = 0
        elif tag == "kml":
            self.write_to_excel()
            return

    def characters(self, content):
        if self.in_placemark == 0 or content.strip() == "":
            return
        elif self.content == "name":
            self.name = content
        elif self.content == "coordinates":
            self.site_coordinate = content
        else:
            return


if __name__ == "__main__":
    parser = xml.sax.make_parser()

    Handler = KmlHandler()
    parser.setContentHandler(Handler)

    parser.parse("test.xml")
